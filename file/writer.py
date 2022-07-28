"""
author: Jiayi Zhu
"""
from enum import Enum

import pandas
from openpyxl import *


class Mode(Enum):
    row = 0
    col = 1


def write_excel(data, filepath: str,
                sheet_name='Sheet1', sheet_index=None,
                header=None, mode=Mode.row):
    """
    将数据写入excel中

    :param data:待写入数据
    :param filepath:文件路径
    :param sheet_name:表单名称
    :param sheet_index:表单序号
    :param header:标题栏
    :param mode: Mode.row: （默认）按行写入；Mode.col: 按列写入
    """
    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.create_sheet(sheet_name, sheet_index)

    t = 1
    if header is not None:
        ws.append(header)
        t = 2
    if mode == Mode.row:
        for r in data:
            ws.append(r)
    if mode == Mode.col:
        for i in range(len(data)):
            for j in range(len(data[i])):
                ws.cell(j + t, i + 1, data[i][j])
    wb.save(filepath)
    wb.close()


if __name__ == '__main__':
    p = r"a.xlsx"
    d = [
        [2, 3, 4, 5],
        ['a', 'b', 'c', 'd']
    ]
    header_ = ['A', 'B']
    write_excel(d, p, 's1', header=header_, mode=Mode.col)
    c = pandas.read_excel(p, sheet_name='s1')
    print(c)
